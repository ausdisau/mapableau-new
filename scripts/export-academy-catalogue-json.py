#!/usr/bin/env python3
"""Export the Academy xlsx workbook to data/academy/catalogue-workbook.json."""

from __future__ import annotations

import json
import sys
from pathlib import Path

try:
    from openpyxl import load_workbook
except ImportError:
    print("openpyxl is required: pip install openpyxl", file=sys.stderr)
    sys.exit(1)

ROOT = Path(__file__).resolve().parents[1]
DEFAULT_XLSX = ROOT / "docs/academy/MapAble_Academy_Master_Course_Catalogue.xlsx"
DEFAULT_JSON = ROOT / "data/academy/catalogue-workbook.json"


def sheet_objs(ws, header_row: int = 3):
    headers = [c.value for c in ws[header_row]]
    rows = []
    for r in range(header_row + 1, ws.max_row + 1):
        vals = [ws.cell(r, c).value for c in range(1, len(headers) + 1)]
        if vals[0] is None:
            continue
        rows.append(
            {
                str(headers[i]): ("" if vals[i] is None else vals[i])
                for i in range(len(headers))
            }
        )
    return rows


def main() -> None:
    xlsx = Path(sys.argv[1]) if len(sys.argv) > 1 else DEFAULT_XLSX
    out = Path(sys.argv[2]) if len(sys.argv) > 2 else DEFAULT_JSON
    wb = load_workbook(xlsx, data_only=True)
    payload = {
        "pathways": sheet_objs(wb["Pathways"]),
        "cursorSeed": sheet_objs(wb["Cursor Seed"]),
        "courseCatalogue": sheet_objs(wb["Course Catalogue"]),
    }
    out.parent.mkdir(parents=True, exist_ok=True)
    out.write_text(json.dumps(payload, ensure_ascii=False, indent=2, default=str) + "\n")
    print(f"Wrote {out} ({len(payload['cursorSeed'])} courses)")


if __name__ == "__main__":
    main()
